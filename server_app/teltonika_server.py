"""
Teltonika GPS Server Module
Protocol-accurate implementation based on DTT v3.2.0 reference.
Supports Codec 8, 8E, 12, 13 over TCP and UDP.
"""

import socket
import threading
import time
import struct
import datetime
import select
import json
import os
from collections import defaultdict


# ─── IO element name registry ──────────────────────────────────────────────────
# Fallback built-in names (small subset).  The full set is loaded from an Excel.
_BUILTIN_IO_NAMES: dict[int, str] = {
    1: "Din1", 2: "Din2", 3: "Din3", 4: "Din4",
    9: "AnalogInput1", 10: "AnalogInput2",
    16: "TotalOdometer", 21: "GsmSignal", 24: "Speed",
    66: "ExternalVoltage", 67: "BatteryVoltage", 68: "BatteryCurrent",
    69: "GnssStatus", 113: "BatteryLevel",
    179: "Dout1", 180: "Dout2", 239: "Ignition", 240: "Movement",
}

# Mutable registry – updated at runtime by load_avl_ids_from_excel()
IO_ELEMENT_NAMES: dict[int, str] = dict(_BUILTIN_IO_NAMES)


def load_avl_ids_from_excel(path: str) -> tuple[dict[int, str], str | None]:
    """Load IO element names from FMB_AVL_IDS.xlsx (MainTable sheet).

    Returns (id→name dict, error_string|None).
    The dict is keyed by column A (Property ID) with the *first* name found per ID.
    """
    import os
    if not os.path.isfile(path):
        return {}, f"File not found: {path}"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if 'MainTable' in wb.sheetnames:
            ws = wb['MainTable']
        else:
            ws = wb.active
        names: dict[int, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            avl_id = row[0]
            prop_name = row[1]
            if avl_id is None or prop_name is None:
                continue
            try:
                avl_id = int(avl_id)
            except (ValueError, TypeError):
                continue
            if avl_id not in names:
                names[avl_id] = str(prop_name).strip()
        wb.close()
        return names, None
    except Exception as e:
        return {}, str(e)


def refresh_io_names(path: str) -> str | None:
    """Reload IO_ELEMENT_NAMES from Excel.  Returns error string or None."""
    names, err = load_avl_ids_from_excel(path)
    if err:
        return err
    # Merge in-place so all existing references stay valid
    IO_ELEMENT_NAMES.clear()
    IO_ELEMENT_NAMES.update(_BUILTIN_IO_NAMES)
    IO_ELEMENT_NAMES.update(names)  # Excel wins over builtins
    return None


def io_name(io_id: int) -> str:
    """Human-readable name for an IO element ID."""
    return IO_ELEMENT_NAMES.get(io_id, f"IO_{io_id}")


# ═══════════════════════════════════════════════════════════════════════════════
#  Hex Packet Annotator
# ═══════════════════════════════════════════════════════════════════════════════

# Color palette for annotations
_C = {
    'preamble': '#546E7A', 'length': '#78909C', 'codec': '#FF8F00',
    'count': '#8E24AA', 'ack': '#2E7D32', 'crc': '#C62828',
    'imei': '#1565C0', 'timestamp': '#F9A825', 'priority': '#E65100',
    'lon': '#00838F', 'lat': '#00695C', 'alt': '#558B2F',
    'angle': '#9E9D24', 'sat': '#00ACC1', 'speed': '#D84315',
    'io_evt': '#6A1B9A', 'io_cnt': '#4A148C',
    'cmd_type': '#00897B', 'cmd_data': '#0277BD',
}

_IO_COLORS = [
    '#00897B', '#5E35B1', '#C0CA33', '#F4511E', '#3949AB',
    '#43A047', '#E53935', '#1E88E5', '#8E24AA', '#FB8C00',
    '#00ACC1', '#7CB342', '#D81B60', '#039BE5', '#FDD835',
]


def _annotate_avl(data: bytes, off: int, n: int, codec_id: int,
                  count: int, sections: list, add, get_io_color) -> int:
    """Annotate AVL records. Returns updated offset."""
    is_8e = (codec_id == 0x8E)
    id_sz = 2 if is_8e else 1
    cnt_sz = 2 if is_8e else 1

    def ru(size):
        nonlocal off
        if off + size > n:
            raise IndexError
        val = int.from_bytes(data[off:off + size], 'big', signed=False)
        off += size
        return val

    try:
        for ri in range(count):
            p = f'R{ri + 1} '

            # Timestamp (8 B)
            ts_s = off
            ts_ms = ru(8)
            try:
                dt = datetime.datetime.utcfromtimestamp(ts_ms / 1000.0)
                tstr = dt.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                tstr = str(ts_ms)
            add(ts_s, off, f'{p}Time: {tstr}', _C['timestamp'])

            # Priority (1 B)
            ps = off; prio = ru(1)
            add(ps, off, f'{p}Priority: {prio}', _C['priority'])

            # GPS (15 B)
            g = off
            lon = int.from_bytes(data[off:off + 4], 'big', signed=True) / 1e7; off += 4
            lat = int.from_bytes(data[off:off + 4], 'big', signed=True) / 1e7; off += 4
            alt = int.from_bytes(data[off:off + 2], 'big', signed=True); off += 2
            ang = ru(2); sat = ru(1); spd = ru(2)
            add(g, g + 4, f'{p}Lon: {lon:.7f}', _C['lon'])
            add(g + 4, g + 8, f'{p}Lat: {lat:.7f}', _C['lat'])
            add(g + 8, g + 10, f'{p}Alt: {alt}m', _C['alt'])
            add(g + 10, g + 12, f'{p}Angle: {ang}°', _C['angle'])
            add(g + 12, g + 13, f'{p}Sats: {sat}', _C['sat'])
            add(g + 13, g + 15, f'{p}Speed: {spd}', _C['speed'])

            # IO Event + Total
            es = off; evio = ru(id_sz)
            add(es, off, f'{p}EventIO: {io_name(evio)} ({evio})', _C['io_evt'])
            ts2 = off; tio = ru(cnt_sz)
            add(ts2, off, f'{p}TotalIO: {tio}', _C['io_cnt'])

            # IO sections: 1B, 2B, 4B, 8B
            for vs in [1, 2, 4, 8]:
                cs = off; ic = ru(cnt_sz)
                add(cs, off, f'{p}{vs}B Count: {ic}', _C['io_cnt'])
                for _ in range(ic):
                    ids = off; iid = ru(id_sz)
                    vs_s = off; ival = ru(vs)
                    c = get_io_color(iid)
                    nm = io_name(iid)
                    add(ids, vs_s, f'{nm} ID ({iid})', c)
                    add(vs_s, off, f'{nm} = {ival}', c)

            # NX variable-length (Codec 8E)
            if is_8e:
                nxs = off; nxc = ru(cnt_sz)
                add(nxs, off, f'{p}NX Count: {nxc}', _C['io_cnt'])
                for _ in range(nxc):
                    ids = off; iid = ru(id_sz)
                    ls = off; vl = ru(2)
                    c = get_io_color(iid)
                    nm = io_name(iid)
                    add(ids, ls, f'{nm} ID ({iid})', c)
                    add(ls, off, f'{nm} Len ({vl})', c)
                    ds = off
                    off += vl
                    if off <= n:
                        add(ds, off, f'{nm} Data', c)

    except (IndexError, struct.error):
        pass
    return off


def annotate_packet(raw_hex: str, protocol: str = 'TCP') -> list[dict]:
    """Annotate raw hex packet with labeled byte ranges.

    Returns list of dicts: {s, e, label, color}  (byte indices, exclusive end).
    """
    try:
        data = bytes.fromhex(raw_hex)
    except ValueError:
        return [{'s': 0, 'e': len(raw_hex) // 2, 'label': 'Invalid hex', 'color': '#EF5350'}]

    n = len(data)
    if n == 0:
        return []

    sections: list[dict] = []

    def add(s, e, label, color):
        if 0 <= s < e <= n:
            sections.append({'s': s, 'e': e, 'label': label, 'color': color})

    io_color_cache: dict[int, str] = {}
    io_idx = [0]

    def get_io_color(io_id):
        if io_id not in io_color_cache:
            io_color_cache[io_id] = _IO_COLORS[io_idx[0] % len(_IO_COLORS)]
            io_idx[0] += 1
        return io_color_cache[io_id]

    try:
        # ── TCP-specific short packets ─────────────────────────────────
        if protocol == 'TCP':
            if n >= 17 and data[0:2] == b'\x00\x0F':
                add(0, 2, 'IMEI Len (15)', _C['length'])
                imei = data[2:17].decode('ascii', errors='replace')
                add(2, 17, f'IMEI: {imei}', _C['imei'])
                return sections
            if n == 1 and data[0] == 0x01:
                add(0, 1, 'IMEI Accept (0x01)', _C['ack'])
                return sections
            if n == 4:
                cnt = struct.unpack('!I', data)[0]
                if cnt < 256:
                    add(0, 4, f'Data ACK (N={cnt})', _C['ack'])
                    return sections

        # ── UDP-specific packets ───────────────────────────────────────
        if protocol == 'UDP':
            if n == 7 and data[0:2] == b'\x00\x05':
                add(0, 2, 'Resp Length', _C['length'])
                pid = struct.unpack('<H', data[2:4])[0]
                add(2, 4, f'PktId ({pid})', _C['imei'])
                add(4, 5, f'NotUsable', _C['preamble'])
                add(5, 6, f'AvlPktId', _C['count'])
                add(6, 7, f'N={data[6]}', _C['ack'])
                return sections

            if n >= 10 and data[0:4] != b'\x00\x00\x00\x00':
                pl = struct.unpack('!H', data[0:2])[0]
                add(0, 2, f'Length ({pl})', _C['length'])
                pid = struct.unpack('!H', data[2:4])[0]
                add(2, 4, f'PktId ({pid})', _C['imei'])
                add(4, 5, 'NotUsable', _C['preamble'])
                add(5, 6, 'AvlPktId', _C['count'])
                il = struct.unpack('!H', data[6:8])[0]
                add(6, 8, f'IMEI Len ({il})', _C['length'])
                ie = 8 + il
                if ie <= n:
                    imei = data[8:ie].decode('ascii', errors='replace')
                    add(8, ie, f'IMEI: {imei}', _C['imei'])
                off = ie
                if off < n:
                    cid = data[off]
                    add(off, off + 1, f'Codec 0x{cid:02X}', _C['codec'])
                    off += 1
                    if off < n:
                        c1 = data[off]
                        add(off, off + 1, f'Count ({c1})', _C['count'])
                        off += 1
                        off = _annotate_avl(data, off, n, cid, c1,
                                            sections, add, get_io_color)
                        if off < n:
                            add(off, off + 1, 'Count2', _C['count'])
                            off += 1
                if off < n:
                    add(off, n, 'Extra', '#9E9E9E')
                return sections

        # ── Framed packet (TCP or UDP command-response wrapper) ────────
        if n >= 12 and struct.unpack('!I', data[0:4])[0] == 0:
            dl = struct.unpack('!I', data[4:8])[0]
            add(0, 4, 'Preamble', _C['preamble'])
            add(4, 8, f'DataLen ({dl})', _C['length'])

            if 8 < n:
                cid = data[8]
                add(8, 9, f'Codec 0x{cid:02X}', _C['codec'])

                if cid in (0x0C, 0x0D):
                    off = 9
                    if off < n:
                        add(off, off + 1, f'Qty ({data[off]})', _C['count']); off += 1
                    if off < n:
                        ct = data[off]
                        tn = {0x05: 'Command', 0x06: 'Response'}.get(ct, f'0x{ct:02X}')
                        add(off, off + 1, tn, _C['cmd_type']); off += 1
                    if off + 4 <= n:
                        cl = struct.unpack('!I', data[off:off + 4])[0]
                        add(off, off + 4, f'CmdLen ({cl})', _C['length']); off += 4
                        ce = min(off + cl, n)
                        if off < ce:
                            try:
                                txt = data[off:ce].decode('ascii', errors='replace')
                                lbl = txt[:50] + ('…' if len(txt) > 50 else '')
                            except Exception:
                                lbl = 'Data'
                            add(off, ce, lbl, _C['cmd_data']); off = ce
                    if off < n:
                        add(off, off + 1, f'Qty2', _C['count']); off += 1

                elif cid in (0x08, 0x8E, 0x10):
                    off = 9
                    if off < n:
                        c1 = data[off]
                        add(off, off + 1, f'Count ({c1})', _C['count']); off += 1
                        eod = min(8 + dl, n)
                        off = _annotate_avl(data, off, eod, cid, c1,
                                            sections, add, get_io_color)
                        if off < eod:
                            add(off, off + 1, 'Count2', _C['count']); off += 1

            crc_s = 8 + dl
            if crc_s + 4 <= n:
                cv = struct.unpack('!I', data[crc_s:crc_s + 4])[0]
                add(crc_s, crc_s + 4, f'CRC (0x{cv:04X})', _C['crc'])
            total = 8 + dl + 4
            if total < n:
                add(total, n, 'Extra', '#9E9E9E')
            return sections

    except Exception:
        pass

    if not sections:
        add(0, n, 'Raw Data', '#9E9E9E')
    return sections


# ═══════════════════════════════════════════════════════════════════════════════
#  Protocol parser
# ═══════════════════════════════════════════════════════════════════════════════

class TeltonikaProtocol:
    """Complete Teltonika Protocol Parser – Codec 8 / 8E / 12 / 13."""

    CODEC_8  = 0x08
    CODEC_8E = 0x8E
    CODEC_12 = 0x0C
    CODEC_13 = 0x0D
    CODEC_16 = 0x10

    @staticmethod
    def crc16(data: bytes) -> int:
        crc = 0
        poly = 0xA001
        for byte in data:
            crc ^= byte
            for _ in range(8):
                if crc & 1:
                    crc = (crc >> 1) ^ poly
                else:
                    crc >>= 1
        return crc

    # ── IMEI / ping ────────────────────────────────────────────────────────────
    @staticmethod
    def parse_imei_packet(data: bytes):
        """Return IMEI string if *data* is valid 17-byte IMEI packet, else None."""
        if len(data) >= 17 and data[0:2] == b'\x00\x0F':
            return data[2:17].decode('ascii', errors='ignore')
        return None

    @staticmethod
    def is_ping(data: bytes) -> bool:
        return len(data) == 1 and data[0] == 0xFF

    # ── TCP frame parsing ─────────────────────────────────────────────────────
    @staticmethod
    def parse_tcp_data_packet(packet: bytes) -> dict | None:
        """Parse a framed TCP data packet.

        Returns dict with type='data' | 'response' | 'crc_error' or None.
        """
        if len(packet) < 12:
            return None
        preamble = struct.unpack('!I', packet[0:4])[0]
        data_len = struct.unpack('!I', packet[4:8])[0]
        if preamble != 0 or data_len == 0 or data_len > 65535:
            return None
        total = 8 + data_len + 4
        if len(packet) < total:
            return None

        codec_id = packet[8]

        # CRC validation
        payload_for_crc = packet[8:8 + data_len]
        recv_crc = struct.unpack('!I', packet[8 + data_len:total])[0]
        calc_crc = TeltonikaProtocol.crc16(payload_for_crc)
        if recv_crc != calc_crc:
            return {'type': 'crc_error', 'raw': packet.hex().upper()}

        # Codec 12 / 13 – command response
        if codec_id in (TeltonikaProtocol.CODEC_12, TeltonikaProtocol.CODEC_13):
            cmd_type = packet[10]
            if cmd_type == 0x06:  # response
                resp_len = struct.unpack('!I', packet[11:15])[0]
                resp_data = packet[15:15 + resp_len].decode('ascii', errors='ignore')
                return {'type': 'response', 'codec': codec_id,
                        'response': resp_data, 'raw': packet.hex().upper()}
            return {'type': 'codec12_other', 'raw': packet.hex().upper()}

        # Codec 8 / 8E / 16 – AVL data
        if codec_id in (TeltonikaProtocol.CODEC_8, TeltonikaProtocol.CODEC_8E,
                        TeltonikaProtocol.CODEC_16):
            count1 = packet[9]
            avl_bytes = packet[10:8 + data_len - 1]
            records = TeltonikaProtocol.decode_avl_records(codec_id, avl_bytes, count1)
            return {'type': 'data', 'codec': codec_id,
                    'codec_name': f'0x{codec_id:02X}',
                    'count': count1, 'records': records,
                    'raw': packet.hex().upper()}

        return None

    # ── UDP frame parsing ─────────────────────────────────────────────────────
    @staticmethod
    def parse_udp_packet(packet: bytes) -> dict | None:
        """Parse a UDP datagram.

        Frame: Length(2) PacketId(2) NotUsable(1) AvlPacketId(1)
               ImeiLen(2) IMEI(N) CodecId(1) Nod1(1) AVL… Nod2(1)

        If preamble is 00000000 → treat as command response (Codec 12).
        """
        if len(packet) < 10:
            return None

        # Command response inside UDP
        if packet[0:4] == b'\x00\x00\x00\x00':
            return TeltonikaProtocol.parse_tcp_data_packet(packet)

        pkt_len    = struct.unpack('!H', packet[0:2])[0]
        pkt_id     = struct.unpack('!H', packet[2:4])[0]
        not_usable = packet[4]
        avl_pkt_id = packet[5]
        imei_len   = struct.unpack('!H', packet[6:8])[0]

        if 8 + imei_len > len(packet):
            return None

        imei = packet[8:8 + imei_len].decode('ascii', errors='ignore')
        off = 8 + imei_len

        if off + 2 > len(packet):
            return None

        codec_id = packet[off]; off += 1
        count1   = packet[off]; off += 1

        avl_bytes = packet[off:len(packet) - 1]  # up to Nod2
        records = TeltonikaProtocol.decode_avl_records(codec_id, avl_bytes, count1)

        return {
            'type': 'data',
            'udp_pkt_id': pkt_id,
            'udp_not_usable': not_usable,
            'udp_avl_pkt_id': avl_pkt_id,
            'imei': imei,
            'codec': codec_id,
            'codec_name': f'0x{codec_id:02X}',
            'count': count1,
            'records': records,
            'raw': packet.hex().upper(),
        }

    # ── AVL record decoding ───────────────────────────────────────────────────
    @staticmethod
    def decode_avl_records(codec: int, data: bytes, count: int) -> list:
        records = []
        off = 0
        is_8e = (codec == TeltonikaProtocol.CODEC_8E)

        def ru(size):
            nonlocal off
            if off + size > len(data):
                raise IndexError
            val = int.from_bytes(data[off:off + size], 'big', signed=False)
            off += size
            return val

        def rs(size):
            nonlocal off
            if off + size > len(data):
                raise IndexError
            val = int.from_bytes(data[off:off + size], 'big', signed=True)
            off += size
            return val

        try:
            for _ in range(count):
                rec = {}
                # Timestamp 8 B (ms since epoch)
                ts_ms = ru(8)
                dt = datetime.datetime.utcfromtimestamp(ts_ms / 1000.0)
                rec['Timestamp'] = dt.strftime('%Y-%m-%d %H:%M:%S')
                rec['Timestamp_ms'] = ts_ms

                # Priority 1 B
                rec['Priority'] = ru(1)

                # GPS 15 B
                rec['Longitude']  = rs(4) / 10_000_000.0
                rec['Latitude']   = rs(4) / 10_000_000.0
                rec['Altitude']   = rs(2)
                rec['Angle']      = ru(2)
                rec['Satellites'] = ru(1)
                rec['Speed']      = ru(2)

                # IO elements
                id_sz  = 2 if is_8e else 1
                cnt_sz = 2 if is_8e else 1

                event_io = ru(id_sz)
                total_io = ru(cnt_sz)
                rec['Event_IO'] = event_io
                rec['Total_IO'] = total_io

                io_data = {}

                def read_io_section(val_size):
                    n = ru(cnt_sz)
                    for _ in range(n):
                        io_id  = ru(id_sz)
                        io_val = ru(val_size)
                        io_data[io_id] = io_val

                read_io_section(1)
                read_io_section(2)
                read_io_section(4)
                read_io_section(8)

                # Codec 8E NX variable-length elements
                if is_8e:
                    nx_count = ru(cnt_sz)
                    for _ in range(nx_count):
                        io_id  = ru(id_sz)
                        val_ln = ru(2)
                        if off + val_ln > len(data):
                            raise IndexError
                        io_data[io_id] = data[off:off + val_ln].hex().upper()
                        off += val_ln

                rec['IO_Data'] = io_data
                records.append(rec)
        except (IndexError, struct.error):
            pass  # partial decode is fine

        return records

    # ── Codec 12 command building ─────────────────────────────────────────────
    @staticmethod
    def build_codec12_command(cmd_text: str) -> bytes:
        """Build a Codec 12 command packet (TCP-framed)."""
        cmd_bytes = cmd_text.encode('ascii')
        payload = bytearray()
        payload.append(TeltonikaProtocol.CODEC_12)  # codec
        payload.append(0x01)   # qty 1
        payload.append(0x05)   # type = command
        payload += struct.pack('!I', len(cmd_bytes))
        payload += cmd_bytes
        payload.append(0x01)   # qty 2

        crc = TeltonikaProtocol.crc16(bytes(payload))

        pkt  = struct.pack('!I', 0)               # preamble
        pkt += struct.pack('!I', len(payload))     # data field length
        pkt += bytes(payload)
        pkt += struct.pack('!I', crc)
        return pkt

    # ── ACK builders ──────────────────────────────────────────────────────────
    @staticmethod
    def build_tcp_data_ack(record_count: int) -> bytes:
        """4-byte TCP data ACK: 00 00 00 {N}."""
        return struct.pack('!I', record_count)

    @staticmethod
    def build_udp_data_ack(pkt_id: int, not_usable: int,
                           avl_pkt_id: int, record_count: int) -> bytes:
        """7-byte UDP data ACK echoing request header fields.
        Matches DTT: length(2) + pktId(2, LE copy) + notUsable(1) + avlPktId(1) + N(1).
        """
        resp = bytearray(7)
        resp[0] = 0x00
        resp[1] = 0x05
        struct.pack_into('!H', resp, 2, pkt_id)   # big-endian echo
        resp[4] = not_usable & 0xFF
        resp[5] = avl_pkt_id & 0xFF
        resp[6] = record_count & 0xFF
        return bytes(resp)


# ═══════════════════════════════════════════════════════════════════════════════
#  Command queue item
# ═══════════════════════════════════════════════════════════════════════════════

class _QueuedCommand:
    __slots__ = ('command', 'packet', 'created', 'sent_time',
                 'retries', 'max_retries', 'response', 'status', 'callback')

    def __init__(self, command: str, callback=None, max_retries: int = 10):
        self.command = command
        self.packet = TeltonikaProtocol.build_codec12_command(command)
        self.created = datetime.datetime.now()
        self.sent_time = None
        self.retries = 0
        self.max_retries = max_retries
        self.response = None
        self.status = 'queued'
        self.callback = callback


# ═══════════════════════════════════════════════════════════════════════════════
#  Server
# ═══════════════════════════════════════════════════════════════════════════════

STATE_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                          'output', 'server_state.json')

class TeltonikaServer:
    """Teltonika GPS Server – TCP **or** UDP on a single port."""

    def __init__(self, port: int = 8000, protocol: str = 'TCP'):
        self.port = port
        self.protocol_mode = protocol.upper()
        self.running = False

        # TCP
        self.tcp_socket = None
        self.tcp_clients = {}
        self.tcp_imei = {}
        self.tcp_buffers = {}

        # UDP
        self.udp_socket = None
        self.udp_clients = {}  # IMEI → (addr, last_seen)

        # Data stores (newest first)
        self.parsed_records  = []
        self.raw_messages    = []
        self.log_messages    = []
        self.command_history = []

        # Command queues
        self.command_queues: dict = defaultdict(list)
        self._active_cmd: dict = {}

        # Scheduled commands
        self.scheduled_commands: dict = defaultdict(list)

        # Interval
        self.interval_stop_flag: dict = {}
        self.interval_last_record: dict = {}

        self.lock = threading.Lock()
        self.data_event = threading.Event()
        self._data_version = 0
        self._saved_version = 0

        # Load persisted state
        self.load_state()

    @property
    def data_version(self) -> int:
        return self._data_version

    # ── Persistence ────────────────────────────────────────────────────────────
    def load_state(self):
        """Load logs/records from disk."""
        if not os.path.exists(STATE_FILE):
            return
        try:
            with open(STATE_FILE, 'r', encoding='utf-8') as f:
                state = json.load(f)
            with self.lock:
                self.parsed_records = state.get('parsed_records', [])
                self.raw_messages = state.get('raw_messages', [])
                self.log_messages = state.get('log_messages', [])
                self.command_history = state.get('command_history', [])
                # Restore scheduled commands if simple strings
                sched = state.get('scheduled_commands', {})
                self.scheduled_commands = defaultdict(list, sched)
            self.log(f"Loaded state from {STATE_FILE}", "INFO")
        except Exception as e:
            self.log(f"Failed to load state: {e}", "ERROR")

    def save_state(self):
        """Save logs/records to disk."""
        os.makedirs(os.path.dirname(STATE_FILE), exist_ok=True)
        try:
            with self.lock:
                state = {
                    'parsed_records': list(self.parsed_records),
                    'raw_messages': list(self.raw_messages),
                    'log_messages': list(self.log_messages),
                    'command_history': list(self.command_history),
                    'scheduled_commands': dict(self.scheduled_commands),
                }
                # Update saved version to avoid redundant saves
                self._saved_version = self._data_version
            
            with open(STATE_FILE, 'w', encoding='utf-8') as f:
                json.dump(state, f, default=str)
                
        except Exception as e:
            print(f"Save state error: {e}")

    def _saver_loop(self):
        """Background thread to save state periodically."""
        while self.running:
            time.sleep(2)
            if self._data_version != self._saved_version:
                self.save_state()

    # ── Logging ────────────────────────────────────────────────────────────────
    def log(self, message: str, msg_type: str = "INFO"):
        ts = datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]
        print(f"[{ts}] [{msg_type}] {message}")
        entry = {'timestamp': ts, 'type': msg_type, 'message': message}
        with self.lock:
            self.log_messages.insert(0, entry)
            if len(self.log_messages) > 1000:
                self.log_messages = self.log_messages[:1000]
            self._data_version += 1
            self.data_event.set()

    def _add_raw(self, direction: str, data: bytes, protocol: str):
        ts = datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3]
        entry = {'timestamp': ts, 'direction': direction, 'protocol': protocol,
                 'hex': data.hex().upper(), 'length': len(data)}
        with self.lock:
            self.raw_messages.insert(0, entry)
            if len(self.raw_messages) > 1000:
                self.raw_messages = self.raw_messages[:1000]
            self._data_version += 1
            self.data_event.set()

    # ── Lifecycle ──────────────────────────────────────────────────────────────
    @staticmethod
    def check_port(port: int, protocol: str = 'TCP') -> str | None:
        """Test if *port* is available.  Returns error string or None."""
        sock_type = socket.SOCK_STREAM if protocol.upper() == 'TCP' else socket.SOCK_DGRAM
        s = socket.socket(socket.AF_INET, sock_type)
        s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        try:
            s.bind(('0.0.0.0', port))
            s.close()
            return None
        except OSError as e:
            s.close()
            return str(e)

    def start(self) -> str | None:
        """Start the server.  Returns error string or None on success."""
        if self.running:
            return None
        # Pre-flight port check
        err = self.check_port(self.port, self.protocol_mode)
        if err:
            # Maybe zombie?
            self.log(f"Port {self.port} busy. Attempting to kill zombie...", "WARN")
            if self.kill_zombie(self.port, self.protocol_mode):
                time.sleep(0.5) # Wait for it to die
                err = self.check_port(self.port, self.protocol_mode)
            
            if err:
                self.log(f"Cannot start: {err}", "ERROR")
                return f"Port {self.port} unavailable: {err}"

        self.running = True
        self._start_error = None
        if self.protocol_mode == 'TCP':
            threading.Thread(target=self._tcp_server_loop, daemon=True).start()
        else:
            threading.Thread(target=self._udp_server_loop, daemon=True).start()
        threading.Thread(target=self._command_sender_loop, daemon=True).start()
        threading.Thread(target=self._saver_loop, daemon=True).start()
        self.log(f"Server started – {self.protocol_mode} on port {self.port}", "START")
        return None

    def kill_zombie(self, port: int, protocol: str) -> bool:
        """Send a magic kill packet to localhost:port to terminate old instance."""
        try:
            if protocol.upper() == 'TCP':
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.settimeout(1)
                    s.connect(('127.0.0.1', port))
                    s.sendall(b'SERVER_DIE_NOW_PLEASE')
            else:
                with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
                    s.sendto(b'SERVER_DIE_NOW_PLEASE', ('127.0.0.1', port))
            return True
        except Exception:
            return False

    def stop(self):
        self.running = False
        self.save_state()  # Save on stop
        for s in [self.tcp_socket, self.udp_socket]:
            if s:
                try: s.close()
                except: pass
        with self.lock:
            for s in list(self.tcp_clients.keys()):
                try: s.close()
                except: pass
            self.tcp_clients.clear()
            self.tcp_imei.clear()
            self.tcp_buffers.clear()
        self.log("Server stopped", "STOP")

    # ═══════════════════════════════════════════════════════════════════════════
    #  TCP
    # ═══════════════════════════════════════════════════════════════════════════
    def _tcp_server_loop(self):
        try:
            self.tcp_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.tcp_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            self.tcp_socket.bind(('0.0.0.0', self.port))
            self.tcp_socket.listen(10)
            self.tcp_socket.setblocking(False)
            self.log(f"TCP listening on :{self.port}", "START")

            while self.running:
                try:
                    rlist = [self.tcp_socket] + list(self.tcp_clients.keys())
                    readable, _, exceptional = select.select(rlist, [], rlist, 1.0)

                    for s in readable:
                        if s is self.tcp_socket:
                            try:
                                client, addr = self.tcp_socket.accept()
                                client.setblocking(False)
                                with self.lock:
                                    self.tcp_clients[client] = addr
                                    self.tcp_buffers[client] = b''
                                self.log(f"TCP connect {addr[0]}:{addr[1]}", "CONN")
                            except:
                                pass
                        else:
                            try:
                                chunk = s.recv(8192)
                                if chunk:
                                    self._add_raw("RX", chunk, "TCP")
                                    with self.lock:
                                        self.tcp_buffers[s] = self.tcp_buffers.get(s, b'') + chunk
                                    self._process_tcp_buffer(s)
                                else:
                                    self._close_tcp(s)
                            except ConnectionResetError:
                                self._close_tcp(s)
                            except Exception as e:
                                self.log(f"TCP client error: {e}", "ERROR")

                    for s in exceptional:
                        self._close_tcp(s)

                except Exception as e:
                    if self.running:
                        self.log(f"TCP loop error: {e}", "ERROR")
                    time.sleep(0.1)
        except Exception as e:
            self.log(f"TCP bind error: {e}", "ERROR")

    def _process_tcp_buffer(self, sock):
        with self.lock:
            buf = self.tcp_buffers.get(sock, b'')

        while len(buf) > 0:
            # Magic kill packet for stale instance shutdown
            if b'SERVER_DIE_NOW_PLEASE' in buf:
                self.log("Received MAGIC KILL signal via TCP. Shutting down.", "STOP")
                self.stop()
                return

            # Ping
            if buf[0] == 0xFF:
                buf = buf[1:]
                continue

            # IMEI handshake (17 bytes: 00 0F + 15 ASCII)
            if len(buf) >= 2 and buf[0:2] == b'\x00\x0F':
                if len(buf) < 17:
                    break
                imei = TeltonikaProtocol.parse_imei_packet(buf[:17])
                buf = buf[17:]
                if imei:
                    with self.lock:
                        self.tcp_imei[sock] = imei
                    self.log(f"IMEI: {imei}", "IMEI")
                    try:
                        sock.send(b'\x01')
                        self._add_raw("TX", b'\x01', "TCP")
                        self.log("IMEI ACK (0x01)", "ACK")
                    except:
                        self._close_tcp(sock)
                        return
                continue

            # Data / Command response frame
            if len(buf) >= 8:
                preamble = struct.unpack('!I', buf[0:4])[0]
                data_len = struct.unpack('!I', buf[4:8])[0]

                if preamble != 0 or data_len == 0 or data_len > 65535:
                    buf = buf[1:]
                    continue

                total = 8 + data_len + 4
                if len(buf) < total:
                    break

                pkt = buf[:total]
                buf = buf[total:]
                info = TeltonikaProtocol.parse_tcp_data_packet(pkt)

                if info is None or info.get('type') == 'crc_error':
                    self.log("CRC error – packet dropped", "ERROR")
                    continue

                imei = self.tcp_imei.get(sock, 'Unknown')

                if info['type'] == 'data':
                    count = info['count']
                    self.log(f"Data: {count} records ({info['codec_name']}) from {imei}", "DATA")

                    # ACK immediately (before heavy processing, like DTT)
                    ack = TeltonikaProtocol.build_tcp_data_ack(count)
                    try:
                        sock.send(ack)
                        self._add_raw("TX", ack, "TCP")
                        self.log(f"Data ACK ({count})", "ACK")
                    except:
                        self._close_tcp(sock)
                        with self.lock:
                            self.tcp_buffers[sock] = buf
                        return

                    # Store records
                    with self.lock:
                        for rec in info.get('records', []):
                            rec['IMEI'] = imei
                            rec['Protocol'] = 'TCP'
                            self.parsed_records.insert(0, rec)
                        if len(self.parsed_records) > 2000:
                            self.parsed_records = self.parsed_records[:2000]
                        if imei != 'Unknown':
                            self.interval_last_record[imei] = datetime.datetime.now()
                        self._data_version += 1
                        self.data_event.set()

                    # Enqueue any scheduled commands
                    self._enqueue_scheduled(imei)

                    # Send queued command right after ACK
                    self._try_send_queued_command_tcp(sock, imei)

                elif info['type'] == 'response':
                    resp = info['response']
                    self.log(f"Response from {imei}: {resp}", "RESP")
                    self._handle_command_response(imei, resp, 'TCP')

                continue

            # Not enough data
            if len(buf) > 16384:
                self.log("Buffer overflow – clearing", "ERROR")
                buf = b''
            break

        with self.lock:
            self.tcp_buffers[sock] = buf

    def _close_tcp(self, client):
        with self.lock:
            imei = self.tcp_imei.pop(client, None)
            self.tcp_clients.pop(client, None)
            self.tcp_buffers.pop(client, None)
        try:
            client.close()
        except:
            pass
        self.log(f"TCP disconnect (IMEI={imei or '?'})", "DISC")

    # ═══════════════════════════════════════════════════════════════════════════
    #  UDP
    # ═══════════════════════════════════════════════════════════════════════════
    def _udp_server_loop(self):
        try:
            self.udp_socket = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            self.udp_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            self.udp_socket.bind(('0.0.0.0', self.port))
            self.udp_socket.setblocking(False)
            self.log(f"UDP listening on :{self.port}", "START")

            while self.running:
                try:
                    readable, _, _ = select.select([self.udp_socket], [], [], 1.0)
                    if not readable:
                        continue

                    data, addr = self.udp_socket.recvfrom(65535)

                    # Magic kill packet for stale instance shutdown
                    if data == b'SERVER_DIE_NOW_PLEASE':
                        self.log("Received MAGIC KILL signal via UDP. Shutting down.", "STOP")
                        self.stop()
                        return

                    self._add_raw("RX", data, "UDP")

                    info = TeltonikaProtocol.parse_udp_packet(data)
                    if info is None:
                        self.log(f"UDP: unparseable from {addr}", "ERROR")
                        continue

                    if info['type'] == 'data':
                        imei  = info['imei']
                        count = info['count']
                        self.log(f"UDP data: {count} records ({info['codec_name']}) "
                                 f"from {imei} @ {addr[0]}:{addr[1]}", "DATA")

                        # ACK
                        ack = TeltonikaProtocol.build_udp_data_ack(
                            info['udp_pkt_id'], info['udp_not_usable'],
                            info['udp_avl_pkt_id'], count)
                        self.udp_socket.sendto(ack, addr)
                        self._add_raw("TX", ack, "UDP")
                        self.log(f"UDP ACK (PktId={info['udp_pkt_id']}, N={count})", "ACK")

                        # Store
                        with self.lock:
                            self.udp_clients[imei] = (addr, datetime.datetime.now())
                            for rec in info.get('records', []):
                                rec['IMEI'] = imei
                                rec['Protocol'] = 'UDP'
                                self.parsed_records.insert(0, rec)
                            if len(self.parsed_records) > 2000:
                                self.parsed_records = self.parsed_records[:2000]
                            self.interval_last_record[imei] = datetime.datetime.now()
                            self._data_version += 1
                            self.data_event.set()

                        self._enqueue_scheduled(imei)
                        self._try_send_queued_command_udp(imei)

                    elif info.get('type') == 'response':
                        resp = info['response']
                        imei = None
                        with self.lock:
                            for i, (a, _) in self.udp_clients.items():
                                if a == addr:
                                    imei = i
                                    break
                        if imei:
                            self.log(f"UDP response from {imei}: {resp}", "RESP")
                            self._handle_command_response(imei, resp, 'UDP')

                except Exception:
                    if self.running:
                        pass
                    time.sleep(0.05)

        except Exception as e:
            self.log(f"UDP bind error: {e}", "ERROR")

    # ═══════════════════════════════════════════════════════════════════════════
    #  Command queue & sending
    # ═══════════════════════════════════════════════════════════════════════════
    def queue_command(self, imei: str, command: str, callback=None):
        qc = _QueuedCommand(command, callback=callback)
        with self.lock:
            self.command_queues[imei].append(qc)
        self.log(f"Queued for {imei}: {command}", "CMD")

    def schedule_command(self, imei: str, command: str):
        with self.lock:
            self.scheduled_commands[imei].append(command)
        self.log(f"Scheduled for {imei}: {command}", "SCHEDULE")

    def _enqueue_scheduled(self, imei: str):
        with self.lock:
            cmds = self.scheduled_commands.pop(imei, [])
        for cmd in cmds:
            self.queue_command(imei, cmd)

    def _try_send_queued_command_tcp(self, sock, imei: str):
        with self.lock:
            active = self._active_cmd.get(imei)
            if active and active.status == 'waiting':
                return
            q = self.command_queues.get(imei, [])
            if not q:
                return
            qc = q[0]
        try:
            sock.send(qc.packet)
            self._add_raw("TX", qc.packet, "TCP")
            self.log(f"CMD → {imei}: {qc.command}", "CMD")
            with self.lock:
                qc.sent_time = datetime.datetime.now()
                qc.status = 'waiting'
                self._active_cmd[imei] = qc
        except Exception as e:
            self.log(f"Send failed to {imei}: {e}", "ERROR")

    def _try_send_queued_command_udp(self, imei: str):
        with self.lock:
            active = self._active_cmd.get(imei)
            if active and active.status == 'waiting':
                return
            q = self.command_queues.get(imei, [])
            if not q:
                return
            qc = q[0]
            addr_info = self.udp_clients.get(imei)
        if not addr_info:
            return
        addr = addr_info[0]
        try:
            self.udp_socket.sendto(qc.packet, addr)
            self._add_raw("TX", qc.packet, "UDP")
            self.log(f"CMD → {imei} (UDP): {qc.command}", "CMD")
            with self.lock:
                qc.sent_time = datetime.datetime.now()
                qc.status = 'waiting'
                self._active_cmd[imei] = qc
        except Exception as e:
            self.log(f"UDP send failed to {imei}: {e}", "ERROR")

    def _handle_command_response(self, imei: str, response: str, protocol: str):
        with self.lock:
            qc = self._active_cmd.pop(imei, None)
            cmd_text = qc.command if qc else 'Unknown'
            duration = 0
            if qc and qc.sent_time:
                duration = int((datetime.datetime.now() - qc.sent_time).total_seconds() * 1000)
                qc.status = 'completed'
                qc.response = response
                q = self.command_queues.get(imei, [])
                if q and q[0] is qc:
                    q.pop(0)

            self.command_history.insert(0, {
                'timestamp': datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3],
                'imei': imei, 'command': cmd_text, 'response': response,
                'protocol': protocol, 'duration_ms': duration,
            })
            if len(self.command_history) > 1000:
                self.command_history = self.command_history[:1000]
            self.data_event.set()

        if qc and qc.callback:
            try:
                qc.callback(response)
            except:
                pass

    def _command_sender_loop(self):
        """Background retry / timeout loop (every 3 s)."""
        while self.running:
            time.sleep(3)
            try:
                with self.lock:
                    imeis = list(self._active_cmd.keys())
                for imei in imeis:
                    with self.lock:
                        qc = self._active_cmd.get(imei)
                    if not qc or qc.status != 'waiting':
                        continue
                    if qc.sent_time and (datetime.datetime.now() - qc.sent_time).total_seconds() > 30:
                        qc.retries += 1
                        if qc.retries >= qc.max_retries:
                            self.log(f"Command timeout for {imei}: {qc.command}", "ERROR")
                            with self.lock:
                                qc.status = 'error'
                                self._active_cmd.pop(imei, None)
                                q = self.command_queues.get(imei, [])
                                if q and q[0] is qc:
                                    q.pop(0)
                                self.command_history.insert(0, {
                                    'timestamp': datetime.datetime.now().strftime('%H:%M:%S.%f')[:-3],
                                    'imei': imei, 'command': qc.command,
                                    'response': '⏱ TIMEOUT', 'protocol': self.protocol_mode,
                                    'duration_ms': -1,
                                })
                                self.data_event.set()
                        else:
                            self.log(f"Retry {qc.retries}/{qc.max_retries} for {imei}", "CMD")
                            qc.sent_time = None
                            qc.status = 'queued'
                            with self.lock:
                                self._active_cmd.pop(imei, None)
            except Exception:
                pass

    # ═══════════════════════════════════════════════════════════════════════════
    #  Public API
    # ═══════════════════════════════════════════════════════════════════════════
    def clear_data(self):
        """Clear all stored data and save state."""
        with self.lock:
            self.parsed_records.clear()
            self.raw_messages.clear()
            self.log_messages.clear()
            self.command_history.clear()
            self._data_version += 1
        self.save_state()

    def send_command(self, imei: str, command: str) -> bool:
        """Queue & attempt to send a command. Always returns True (queued)."""
        self.queue_command(imei, command)
        if self.protocol_mode == 'TCP':
            sock = self._find_tcp_socket(imei)
            if sock:
                self._try_send_queued_command_tcp(sock, imei)
        else:
            self._try_send_queued_command_udp(imei)
        return True

    def _find_tcp_socket(self, imei: str):
        with self.lock:
            for s, i in self.tcp_imei.items():
                if i == imei and s in self.tcp_clients:
                    return s
        return None

    def is_device_connected(self, imei: str) -> bool:
        with self.lock:
            for s, i in self.tcp_imei.items():
                if i == imei and s in self.tcp_clients:
                    return True
            info = self.udp_clients.get(imei)
            if info:
                _, last = info
                if (datetime.datetime.now() - last).total_seconds() < 300:
                    return True
                else:
                    del self.udp_clients[imei]
        return False

    def get_connected_devices(self) -> list:
        devices = []
        with self.lock:
            for sock, imei in self.tcp_imei.items():
                if sock in self.tcp_clients:
                    addr = self.tcp_clients[sock]
                    devices.append({
                        'IMEI': imei, 'Protocol': 'TCP',
                        'Address': f"{addr[0]}:{addr[1]}", 'Status': 'Connected',
                    })
            now = datetime.datetime.now()
            stale = []
            for imei, (addr, last) in self.udp_clients.items():
                age = (now - last).total_seconds()
                if age < 300:
                    devices.append({
                        'IMEI': imei, 'Protocol': 'UDP',
                        'Address': f"{addr[0]}:{addr[1]}",
                        'Status': f'Last seen {int(age)}s ago',
                    })
                else:
                    stale.append(imei)
            for imei in stale:
                del self.udp_clients[imei]
        return devices

    def get_queue_status(self, imei: str = None) -> dict:
        with self.lock:
            if imei:
                q = self.command_queues.get(imei, [])
                active = self._active_cmd.get(imei)
                return {
                    'queued': len(q),
                    'active': active.command if active and active.status == 'waiting' else None,
                    'retries': active.retries if active else 0,
                }
            status = {}
            for im, q in self.command_queues.items():
                active = self._active_cmd.get(im)
                if q or active:
                    status[im] = {
                        'queued': len(q),
                        'active': active.command if active and active.status == 'waiting' else None,
                    }
            return status

    def stop_interval_command(self, imei: str):
        with self.lock:
            self.interval_stop_flag[imei] = True
        self.log(f"Stop requested for interval on {imei}", "INFO")

    def send_command_with_interval(self, imei: str, command: str,
                                   interval_sec: float = 0, duration_sec: float = 0,
                                   wait_for_record: bool = False):
        """Repeatedly queue a command at *interval_sec* for *duration_sec*."""
        with self.lock:
            self.interval_stop_flag[imei] = False

        result = {'success': True, 'commands_sent': 0, 'errors': [], 'stopped': False}
        send_once = interval_sec == 0 or duration_sec == 0
        start = datetime.datetime.now()
        end = start + datetime.timedelta(seconds=duration_sec)

        self.log(f"Interval start: '{command}' every {interval_sec}s "
                 f"for {duration_sec}s to {imei}", "INFO")

        try:
            cycle = 0
            while True:
                with self.lock:
                    if self.interval_stop_flag.get(imei):
                        result['stopped'] = True
                        break

                if not send_once and datetime.datetime.now() >= end:
                    break

                cycle += 1

                if wait_for_record:
                    wait_start = datetime.datetime.now()
                    while (datetime.datetime.now() - wait_start).total_seconds() < 120:
                        with self.lock:
                            if self.interval_stop_flag.get(imei):
                                result['stopped'] = True
                                return result
                            lr = self.interval_last_record.get(imei)
                        if lr and lr > wait_start:
                            break
                        time.sleep(0.3)
                    else:
                        result['errors'].append(f"Timeout waiting for record (cycle {cycle})")
                        if send_once:
                            break
                        continue

                ok = self.send_command(imei, command)
                if ok:
                    result['commands_sent'] += 1
                else:
                    result['errors'].append(f"Send failed (cycle {cycle})")

                if send_once:
                    break

                target = cycle * interval_sec
                elapsed = (datetime.datetime.now() - start).total_seconds()
                sl = target - elapsed
                if sl > 0:
                    while sl > 0:
                        time.sleep(min(sl, 0.5))
                        sl -= 0.5
                        with self.lock:
                            if self.interval_stop_flag.get(imei):
                                result['stopped'] = True
                                return result

        except Exception as e:
            result['success'] = False
            result['errors'].append(str(e))

        self.log(f"Interval done: sent {result['commands_sent']} to {imei}", "INFO")
        return result
